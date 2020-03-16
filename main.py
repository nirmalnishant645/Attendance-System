#Modules
from tkinter import *
from tkinter import messagebox as ms
from tkinter import ttk
from tkinter import font
from tkinter import Message, Text
import tkinter as tk
import cv2
import os
import csv
import numpy as np
from PIL import Image, ImageTk
import pandas as pd
import datetime
import time
import sqlite3 as sql
import shutil
from openpyxl import Workbook, load_workbook
import webbrowser
from tempfile import NamedTemporaryFile
import sys
from PyQt5.QtWidgets import QApplication, QWidget, QScrollArea, QTableWidget, QVBoxLayout, QTableWidgetItem
import matplotlib.pyplot as plt

#Make database and users table(if doesn't already exist)
with sql.connect('faculty.db') as db:
    c = db.cursor()

c.execute('CREATE TABLE IF NOT EXISTS user (name TEXT NOT NULL, username TEXT NOT NULL, password TEXT NOT NULL, question TEXT NOT NULL, answer TEXT NOT NULL);')
db.commit()
db.close()

#main class
class main:
    def __init__(self, master):
        #Window
        self.master = master
        #Variables
        self.name = StringVar()
        self.username = StringVar()
        self.password = StringVar()
        self.question = StringVar()
        self.answer = StringVar()
        self.n_name = StringVar()
        self.n_username = StringVar()
        self.n_password = StringVar()
        self.n_question = StringVar()
        self.n_answer = StringVar()
        self.n_level = StringVar()
        self.subject = StringVar()
        self.studID = StringVar()
        self.studName = StringVar()
        self.n_subject = StringVar()
        self.year = StringVar()
        self.n_year = StringVar()
        self.month = StringVar()
        self.n_month = StringVar()
        self.options = ["What primary school did you attend?",
                        "What is the middle name of your father?",
                        "What time of the day were you born?",
                        "What was your childhood nickname?",]

        self.level = open("Course\Level.txt").read().splitlines()
        self.arts = open("Course\Arts.txt").read().splitlines()
        self.science = open("Course\Science.txt").read().splitlines()
        self.courses = {"BA": self.arts, "BSc": self.science, "MA":self.arts, "MSc": self.science}
        self.year = ['First Year', 'Second Year', 'Third Year']
        self.month = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']

        #Widgets
        self.widgets()

    #Login Function
    def login(self):
        #Build Connection to DB
        with sql.connect('faculty.db') as db:
            c= db.cursor()

        #Find user in the database and take proper action
        find_user = ('SELECT * FROM user WHERE username = ? and password = ?')
        c.execute(find_user,[(self.username.get()), (self.password.get())])
        result = c.fetchone()
        if result:
            ms.showinfo(title = 'Success!', message = self.username.get() + ' logged in!')
            self.dashB()
        else:
            ms.showerror(title = 'Error', message = 'Username or Password incorrect!')

    #New User
    def new_user(self):
        #Build Connectio to DB
        with sql.connect('faculty.db') as db:
            c = db.cursor()

        #Find existing username (if any), take proper action
        find_user = ('SELECT * FROM user WHERE username = ?')
        c.execute(find_user,[(self.n_username.get())])
        if c.fetchone():
            ms.showerror(title = 'Error!', message = 'Username already taken!')
        else:
            ms.showinfo(title = 'Success!', message = 'Faculty Registration Completed!')
            self.log()
            #Create new account
            insert = 'INSERT INTO user(name, username, password, question, answer) values(?, ?, ?, ?, ?)'
            c.execute(insert,[(self.n_name.get()),(self.n_username.get()),(self.n_password.get()),(self.n_question.get()),(self.n_answer.get())])
            db.commit()

    #Attendance
    def attendance(self):
        self.n_level.set('')
        self.n_subject.set('')
        self.dashBf.place_forget()
        self.subCombo = ttk.Combobox(self.attendancef, textvariable = self.n_subject, font = ('', 15))
        self.subCombo.grid(row = 1, column = 2)
        self.attendancef.place(relx = 0.5, rely = 0.5, anchor = CENTER)

    #Report
    def report(self):
        self.studID.set('')
        self.studName.set('')
        self.n_level.set('')
        self.n_subject.set('')
        self.n_year.set('')
        self.dashBf.place_forget()
        self.subCombo = ttk.Combobox(self.reportf, textvariable = self.n_subject, font = ('', 15))
        self.subCombo.grid(row = 1, column = 2)
        self.reportf.place(relx = 0.5, rely = 0.5, anchor = CENTER)
    
    # Attendance Sheet
    def class_report(self):
        files = os.listdir('Attendance/')
        course = self.n_level.get() + self.n_subject.get()
        month = self.n_month.get()
        month = self.month.index(month) + 1
        file = '_' + course + '-' + str(month) + '.xlsx'
        if file in files:
            df = pd.read_excel('Attendance/' + file)
            df = df[df[self.month[month - 1]].notna()]
            present = {'Present'}
            df['Total Present'] = df.isin(present).sum(1)
            app = QApplication(sys.argv)
            win = QWidget()
            scroll = QScrollArea()
            layout = QVBoxLayout()
            table = QTableWidget()
            scroll.setWidget(table)
            layout.addWidget(table)
            win.setLayout(layout)
            QWidget.setWindowTitle(win, 'Attendance Sheet')
            date = [str(i) for i in range(33)]
            date[0] = 'Roll No. - Name'
            date[-1] = 'Total Present'
            now = datetime.datetime.now()
            if month == int(now.month):
                ts = time.time()
                date1 = datetime.datetime.fromtimestamp(ts).strftime('%d')
                column_name = {str(i) + '-' + str(month) : 'Absent' for i in range(int(date1) + 1)}
            else:
                column_name = {str(i) + '-' + str(month) : 'Absent' for i in range(32)}
            df.fillna(column_name, inplace = True)
            table.setColumnCount(len(df.columns))
            table.setRowCount(len(df.index))
            table.setHorizontalHeaderLabels(date)
            for i in range(len(df.index)):
                for j in range(len(df.columns)):
                    table.setItem(i, j, QTableWidgetItem(str(df.iloc[i, j])))
            win.showMaximized()
            app.exec_()
        else:
            ms.showerror(title = 'Error!', message = 'Record for selected Course and Month does not exist!')

    # Graph
    def student_report(self):
        files = os.listdir('Attendance/')
        course = self.n_level.get() + self.n_subject.get()
        month = self.n_month.get()
        month = self.month.index(month) + 1
        file = '_' + course + '-' + str(month) + '.xlsx'
        if file in files:
            df = pd.read_excel('Attendance/' + file)
            present = {'Present'}
            df2 = pd.DataFrame()
            df2['Students'] = df[self.n_month.get()]
            df2['Total Present'] = df.isin(present).sum(1)
            #print(df2)
            histo = {}
            for i in range(len(df2.index)):
                if type(df2['Students'][i]) == str:
                    histo[df2['Students'][i]] = df2['Total Present'][i]
            plt.bar(range(len(histo)), list(histo.values()), align = 'center')
            plt.xticks(range(len(histo)), list(histo.keys()), rotation = 'vertical')
            plt.subplots_adjust(left = 0.05, bottom = 0.3, right = 0.95, top = 0.95 )
            labels = list(histo.values())
            for i in range(len(histo)):
                plt.text(x = i, y = int(labels[i]), s = labels[i])
            plt.title('Student Report')
            mng = plt.get_current_fig_manager()
            mng.window.state('zoomed')
            plt.show()
        else:
            ms.showerror(title = 'Error!', message = 'Record for selected Course and Month does not exist!')

    #Student Registration
    def studReg(self):
        self.studID.set('')
        self.studName.set('')
        self.n_level.set('')
        self.n_subject.set('')
        self.n_year.set('')
        self.dashBf.place_forget()
        self.subCombo = ttk.Combobox(self.studRegf, textvariable = self.n_subject, font = ('', 15))
        self.subCombo.grid(row = 1, column = 2)
        self.studRegf.place(relx = 0.5, rely = 0.5, anchor = CENTER)

    def take_images(self):
        cam = cv2.VideoCapture(0)
        id = self.studID.get()
        name = self.studName.get()
        course = self.n_level.get() + self.n_subject.get() + self.n_year.get()
        detector = cv2.CascadeClassifier('data\haarcascade\haarcascade_frontalface_default.xml')
        sampleNum = 0
        row = [course, id , name]
        with open('StudentDetails\StudentDetails.csv','r') as csvFile:
            reader = csv.reader(csvFile, delimiter = '\t')
            for line in reader:
                if line:
                    line = line[0].split(',')
                    if line[0] == row[0] and line[1] == row[1]:
                        ms.showerror(title = 'Error!', message = 'Student with course, '+course+' and ID, '+id+' alrady exist with name, '+line[2])
                        csvFile.close()
                        return
        while True:
            ret, img = cam.read()
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = detector.detectMultiScale(gray, 1.3, 5)
            for (x,y,w,h) in faces:
                cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
                #incrementing sample number
                sampleNum += 1
                #saving the captured face in the dataset folder TrainingImage
                cv2.imwrite("TrainingImage\ "+ name +"."+ id +'.'+ str(sampleNum) + ".jpg", gray[y : y + h, x : x + w])
                #display the frame
                cv2.imshow('frame', img)
            #wait for 100 miliseconds
            if cv2.waitKey(100) & 0xFF == ord('q'):
                break
            # break if the sample number is morethan 100
            elif sampleNum > 60:
                break
        cam.release()
        cv2.destroyAllWindows()
        res = "Images Saved for ID : " + id +" Name : "+ name
        with open('StudentDetails\StudentDetails.csv','a+') as csvFile:
            writer = csv.writer(csvFile)
            writer.writerow(row)
        csvFile.close()
        ms.showinfo(title = 'Done!', message = res)

    def train(self):
        recognizer = cv2.face.LBPHFaceRecognizer_create()
        detector = cv2.CascadeClassifier('data\haarcascade\haarcascade_frontalface_default.xml')

        imagePaths = [os.path.join('TrainingImage', f) for f in os.listdir('TrainingImage')]
        faces = []
        ids = []

        for imagePath in imagePaths:
            if imagePath != 'TrainingImage\\readme.jpg':
                pilImage = Image.open(imagePath).convert('L')
                imageNP = np.array(pilImage, 'uint8')
                id = int(os.path.split(imagePath)[-1].split(".")[1])
                faces.append(imageNP)
                ids.append(id)

        recognizer.train(faces, np.array(ids))
        recognizer.save("TrainingImageLabel\Trainer.yml")
        ms.showinfo(title = 'Done!', message = 'Images Saved Succesfully!')

    def track(self):
        course = self.n_level.get() + self.n_subject.get()
        now = datetime.datetime.now()
        fname = r'Attendance\_' + course + '-' + str(now.month)+ '.xlsx'
        if os.path.isfile(fname):
            book = load_workbook(fname)
        else:
            book = Workbook()
            book.save(fname)
        sheet = book.active
        recognizer = cv2.face.LBPHFaceRecognizer_create()
        recognizer.read("TrainingImageLabel\Trainer.yml")
        faceCascade = cv2.CascadeClassifier('data\haarcascade\haarcascade_frontalface_default.xml')
        df = pd.read_csv("StudentDetails\StudentDetails.csv")
        cam = cv2.VideoCapture(0)
        font = cv2.FONT_HERSHEY_SIMPLEX
        col_names = ['ID', 'NAME', 'DATE', 'TIME']
        attendance = pd.DataFrame(columns = col_names)
        for i in range(1, 32):
            sheet.cell(row = 1, column = i+1).value = str(i) + '-' + str(now.month)
        sheet.cell(row = 1, column = 1).value = self.month[now.month - 1]
        while True:
            ret, im = cam.read()
            gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
            faces = faceCascade.detectMultiScale(gray, 1.2, 5)
            for(x, y, w, h) in faces:
                cv2.rectangle(im, (x,y), (x+w, y+h), (255, 0, 0), 2)
                id, conf = recognizer.predict(gray[y:y+h, x:x+w])
                if conf < 50:
                    ts = time.time()
                    date = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d')
                    timeStamp = datetime.datetime.fromtimestamp(ts).strftime('%H:%M:%S')
                    aa = df.loc[df['ID'] == id]['Name'].values
                    tt = str(id) + "-" + aa
                    stud_name = str(tt)
                    stud_name = stud_name.replace('[', '').replace(']', '')
                    stud_name = stud_name[1:-1]
                    attendance.loc[len(attendance)] = [id, aa, date, timeStamp]
                    sheet.cell(row = int(id) + 1, column = 1).value = str(stud_name)
                    sheet.cell(row = int(id) + 1, column = int(now.day) + 1).value = "Present"

                else:
                    id = 'Unknown'
                    tt = str(id)
                if conf > 75:
                    noOfFile = len(os.listdir("ImagesUnknown")) + 1
                    cv2.imwrite("ImagesUnknown\Image" + str(noOfFile) + ".jpg", im[y:y+h, x:x+w])
                cv2.putText(im, str(tt), (x, y+h), font, 1, (255, 255, 255), 2)
            attendance = attendance.drop_duplicates(subset = ['ID'], keep = 'first')
            cv2.imshow('im', im)
            if (cv2.waitKey(1) == ord('q')):
                break
        book.save(fname)
        cam.release()
        cv2.destroyAllWindows()
        ms.showinfo(title = 'Done!', message = 'Attandance saved succesfully!')



    #Logout
    def logout(self):
        self.log()
        self.dashBf.place_forget()

    #Frame Packing
    def log(self):
        self.username.set('')
        self.password.set('')
        self.regf.place_forget()
        self.logf.place(relx = 0.5, rely = 0.5, anchor = CENTER)

    def reg(self):
        self.n_name.set('')
        self.n_username.set('')
        self.n_password.set('')
        self.n_question.set('')
        self.n_answer.set('')
        self.logf.place_forget()
        self.regf.place(relx = 0.5, rely = 0.5, anchor = CENTER)

    def dashB(self):
        self.attendancef.place_forget()
        self.studRegf.place_forget()
        self.reportf.place_forget()
        self.logf.place_forget()
        with sql.connect('faculty.db') as db:
            c = db.cursor()
        find_name = ('SELECT name FROM user WHERE username = ?')
        c.execute(find_name,[(self.username.get())])
        name = c.fetchone()
        name = name[0].split()
        self.head['text'] = name[0]
        self.dashBf.place(relx = 0.5, rely = 0.5, anchor = CENTER)

    def changeSubject(self, event):
        self.subCombo['values'] = self.courses[self.levelCombo.get()]

    def close(self):
        root.destroy()


    #Draw widgets
    def widgets(self):
        self.background_label = Label(self.master, image = root.image)
        self.background_label.place(x = 0, y = 0, relwidth = 1, relheight = 1)
        Button(self.master, text = 'X', font = ('', 12), command = self.close).place(relx = 0.97, rely = 0.001)

        self.logf = Frame(self.master, highlightbackground = 'black', highlightcolor = 'black', highlightthickness = 3, bg = 'white')
        Label(self.logf, text = 'Login', font = ('', 35), bg = 'white').grid(sticky = W, row = 0, column = 1)
        Label(self.logf, text = '', font = ('', 25), bg = 'white').grid(row = 1, column = 1)
        Label(self.logf, text = 'Username: ', font = ('', 20), pady = 5, padx = 5, bg = 'white').grid(sticky = W, row = 2, column = 0)
        Entry(self.logf, textvariable = self.username, bd = 5, font = ('', 15), bg = 'white', width = 23).grid(sticky = W, row = 2, column = 1)
        Label(self.logf, text = 'Password: ', font = ('', 20), pady = 5, padx = 5, bg = 'white').grid(sticky = W, row = 4, column = 0)
        Entry(self.logf, textvariable = self.password, bd = 5, font = ('', 15), show = '*', bg = 'white', width = 23).grid(sticky = W, row = 4, column = 1)
        Label(self.logf, text = '', font = ('', 5), bg = 'white').grid(row = 4, column = 2)
        Label(self.logf, text = '', font = ('', 5), bg = 'white').grid(row = 5, column = 0)
        Button(self.logf, text = ' Register ', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.reg, bg = 'white').grid(row = 6, column = 0)
        Button(self.logf, text = ' Login ', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.login, bg  = 'white').grid(row = 6, column = 1)
        self.logf.place(relx = 0.5, rely = 0.5, anchor = CENTER)

        self.regf = Frame(self.master,  highlightbackground = 'black', highlightcolor = 'black', highlightthickness = 3, bg = 'white')
        Label(self.regf, text = 'Registration', font = ('', 30), bg = 'white').grid(sticky = W, row = 0, column = 1)
        Label(self.regf, text = '', font = ('', 20), bg = 'white').grid(row = 1, column = 1)
        Label(self.regf, text = 'Name: ', font = ('', 20), pady = 5, padx = 5, bg = 'white').grid(sticky = W, row = 2, column = 0)
        Entry(self.regf, textvariable = self.n_name, bd = 5, font = ('', 15), bg = 'white').grid(row = 2, column = 1)
        Label(self.regf, text = 'Username: ', font = ('', 20), pady = 5, padx = 5, bg = 'white').grid(sticky = W)
        Entry(self.regf, textvariable = self.n_username, bd = 5, font = ('', 15), bg = 'white').grid(row = 3, column = 1)
        Label(self.regf, text = 'Password: ', font = ('', 20), pady = 5, padx = 5, bg = 'white').grid(sticky = W)
        Entry(self.regf, textvariable = self.n_password, bd = 5, font = ('', 15), show = '*', bg = 'white').grid(row = 4, column = 1)
        Label(self.regf, text = 'Security Question: ', font = ('', 20), pady = 5, padx = 5, bg = 'white').grid(sticky = W)
        ttk.Combobox(self.regf, textvariable = self.n_question, values = self.options, font = ('', 15)).grid(row = 5, column = 1)
        Label(self.regf, text = 'Security Answer: ', font = ('', 20), pady = 5, padx = 5, bg = 'white').grid(sticky = W)
        Entry(self.regf, textvariable = self.n_answer, bd = 5, font = ('', 15), show = '*', bg = 'white').grid(row = 6, column = 1)
        Button(self.regf, text = 'Register', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.new_user, bg = 'white').grid(row = 7, column = 1)
        Button(self.regf, text = 'Go to Login', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.log, bg = 'white').grid(sticky = E, row = 7, column = 3)

        self.dashBf = Frame(self.master,  highlightbackground = 'black', highlightcolor = 'black', highlightthickness = 3, bg = 'white')
        Label(self.dashBf, text = 'Welcome', font = ('', 30), bg = 'white').grid(sticky = E, row = 0, column = 3)
        self.head = Label(self.dashBf, text = '', font = ('', 30), bg = 'white')
        self.head.grid(sticky = W, row = 0, column = 5)
        Label(self.dashBf, text = '', padx = 5, pady = 5, bg = 'white').grid(row = 1, column = 4)
        Label(self.dashBf, text = '', padx = 5, pady = 5, bg = 'white').grid(row = 0, column = 4)
        Label(self.dashBf, text = '', font = ('', 35), padx = 5, pady = 5, bg = 'white').grid(row = 2, column = 1)
        Button(self.dashBf, text = 'Attendance', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.attendance, bg = 'white', width = 20).grid(sticky = W, row = 2, column = 3)
        Button(self.dashBf, text = 'View Report', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.report, bg = 'white', width = 20).grid(sticky = E, row = 2, column = 5)
        Label(self.dashBf, text = '', padx = 5, pady = 5, bg = 'white').grid(row = 2, column = 6)
        Label(self.dashBf, text = '', padx = 5, pady = 5, bg = 'white').grid(row = 3, column = 3)
        Label(self.dashBf, text = '', padx = 5, pady = 5, bg = 'white').grid(row = 3, column = 5)
        Label(self.dashBf, text = '', padx = 5, pady = 5, bg = 'white').grid(row = 4, column = 1)
        Button(self.dashBf, text = 'Student Registration', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.studReg, bg = 'white', width = 20).grid(sticky = W, row = 4, column = 3)
        Button(self.dashBf, text = 'Logout', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.logout, bg = 'white', width = 20).grid(sticky = E, row = 4, column = 5)
        Label(self.dashBf, text = '', padx = 5, pady = 5, bg = 'white').grid(row = 4, column = 6)

        self.attendancef = Frame(self.master,  highlightbackground = 'black', highlightcolor = 'black', highlightthickness = 3, bg = 'white')
        Label(self.attendancef, text = 'Attendance', font = ('', 30), bg = 'white').grid(sticky = W, row = 0, column = 1)
        Label(self.attendancef, text = 'Select Course: ', font = ('', 20), pady = 5, padx = 5, bg = 'white').grid(sticky = W, row = 1, column = 0)
        self.levelCombo = ttk.Combobox(self.attendancef, textvariable = self.n_level, values = self.level, font = ('', 15))
        self.levelCombo.bind('<<ComboboxSelected>>', self.changeSubject)
        self.levelCombo.grid(row = 1, column = 1)
        Button(self.attendancef, text = 'Take Attendance', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.track, bg = 'white').grid(row = 2, column = 1)
        Button(self.attendancef, text = 'Back', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.dashB, bg = 'white').grid(row = 2, column = 2)

        self.studRegf = Frame(self.master,  highlightbackground = 'black', highlightcolor = 'black', highlightthickness = 3, bg = 'white')
        Label(self.studRegf, text = 'Student Registration', font = ('', 30), bg = 'white').grid(sticky = W, row = 0, column = 1)
        Label(self.studRegf, text = 'Select Course', font = ('', 20), pady = 5, padx = 5, bg = 'white').grid(sticky = W, row = 1, column = 0)
        self.levelCombo = ttk.Combobox(self.studRegf, textvariable = self.n_level, values = self.level, font = ('', 15))
        self.levelCombo.bind('<<ComboboxSelected>>', self.changeSubject)
        self.levelCombo.grid(row = 1, column = 1)
        Label(self.studRegf, text = 'Year', font = ('', 20), pady = 5, padx = 5, bg = 'white').grid(sticky = W, row = 2, column = 0)
        self.yearCombo = ttk.Combobox(self.studRegf, textvariable = self.n_year, values = self.year, font = ('', 15))
        self.yearCombo.grid(row = 2, column = 1)
        Label(self.studRegf, text = '', pady = 5, padx = 5, bg = 'white').grid(row = 1, column = 3)
        Label(self.studRegf, text = 'ID', font = ('', 20), pady = 5, padx = 5, bg = 'white').grid(sticky = W, row = 3, column = 0)
        Entry(self.studRegf, textvariable = self.studID, bd = 5, font = ('', 15), bg = 'white').grid(row = 3, column = 1)
        Label(self.studRegf, text = 'Name', font = ('', 20), pady = 5, padx = 5, bg = 'white').grid(sticky = W, row = 4, column = 0)
        Entry(self.studRegf, textvariable = self.studName, bd = 5, font = ('', 15), bg = 'white').grid(row = 4, column = 1)
        Button(self.studRegf, text = 'Take Image', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.take_images, bg = 'white').grid(sticky = W, row = 5, column = 1)
        Button(self.studRegf, text = 'Back', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.dashB, bg = 'white', width = 10).grid(sticky = E, row = 5, column = 2)
        Button(self.studRegf, text = 'Save', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.train, bg = 'white', width = 10).grid(sticky = E, row = 5, column = 1)

        self.reportf = Frame(self.master,  highlightbackground = 'black', highlightcolor = 'black', highlightthickness = 3, bg = 'white')
        Label(self.reportf, text = 'Reports', font = ('', 30), bg = 'white').grid(sticky = W, row = 0, column = 1)
        Label(self.reportf, text = 'Select Course', font = ('', 20), pady = 5, padx = 5, bg = 'white').grid(sticky = W, row = 1, column = 0)
        self.levelCombo = ttk.Combobox(self.reportf, textvariable = self.n_level, values = self.level, font = ('', 15))
        self.levelCombo.bind('<<ComboboxSelected>>', self.changeSubject)
        self.levelCombo.grid(row = 1, column = 1)
        Label(self.reportf, text = 'Month', font = ('', 20), pady = 5, padx = 5, bg = 'white').grid(sticky = W, row = 2, column = 0)
        self.yearCombo = ttk.Combobox(self.reportf, textvariable = self.n_month, values = self.month, font = ('', 15))
        self.yearCombo.grid(row = 2, column = 1)
        Label(self.reportf, text = '', pady = 5, padx = 5, bg = 'white').grid(row = 1, column = 3)
        Button(self.reportf, text = 'Attendance Sheet', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.class_report, bg = 'white').grid(sticky = E, row = 5, column = 1)
        Button(self.reportf, text = 'Back', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.dashB, bg = 'white', width = 10).grid(sticky = E, row = 5, column = 0)
        Button(self.reportf, text = 'Student Report', bd = 3, font = ('', 15), padx = 5, pady = 5, command = self.student_report, bg = 'white').grid(sticky = E, row = 5, column = 2)

#Create Window and Application Object
root = Tk()
root.attributes('-fullscreen', True)
root.image = tk.PhotoImage(file = 'GUI/banner.gif')
main(root)
root.mainloop()
